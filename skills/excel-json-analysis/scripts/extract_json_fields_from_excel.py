#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract fields from JSON text stored in an Excel column.",
    )
    parser.add_argument("--input", required=True, help="Path to the source xlsx file.")
    parser.add_argument("--sheet", help="Sheet name. Defaults to the first sheet.")
    parser.add_argument(
        "--json-column",
        required=True,
        help="Header name for the column that stores JSON text.",
    )
    parser.add_argument(
        "--records-path",
        default="",
        help="Dot path to the list/object containing records, e.g. a.b.orderLines.",
    )
    parser.add_argument(
        "--fields",
        nargs="+",
        required=True,
        help="Field names to extract from each record.",
    )
    parser.add_argument(
        "--dedupe-fields",
        nargs="*",
        default=None,
        help="Fields used as the dedupe key. Defaults to all extracted fields.",
    )
    parser.add_argument(
        "--output",
        help="Output xlsx path. Defaults to <input_stem>_extracted.xlsx.",
    )
    return parser.parse_args()


def resolve_records(payload: Any, records_path: str) -> list[dict[str, Any]]:
    current = payload
    if records_path:
        for part in records_path.split("."):
            if not isinstance(current, dict):
                return []
            current = current.get(part)

    if isinstance(current, list):
        return [item for item in current if isinstance(item, dict)]
    if isinstance(current, dict):
        return [current]
    return []


def build_output_path(input_path: Path, output_arg: str | None) -> Path:
    if output_arg:
        return Path(output_arg).expanduser().resolve()
    return input_path.with_name(f"{input_path.stem}_extracted.xlsx")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = build_output_path(input_path, args.output)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet_name = args.sheet or workbook.sheetnames[0]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    if args.json_column not in headers:
        raise ValueError(
            f"JSON column '{args.json_column}' not found. Available headers: {headers}"
        )
    json_index = headers.index(args.json_column)

    dedupe_fields = args.dedupe_fields or args.fields
    unique_rows: OrderedDict[tuple[str, ...], list[str]] = OrderedDict()

    row_count = 0
    json_row_count = 0
    record_count = 0
    invalid_json_count = 0
    missing_records_count = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_count += 1
        payload_text = row[json_index]
        if not payload_text:
            continue
        json_row_count += 1

        try:
            payload = json.loads(payload_text)
        except json.JSONDecodeError:
            invalid_json_count += 1
            continue

        records = resolve_records(payload, args.records_path)
        if not records:
            missing_records_count += 1
            continue

        for record in records:
            values = [str(record.get(field, "")).strip() for field in args.fields]
            dedupe_key = tuple(
                str(record.get(field, "")).strip() for field in dedupe_fields
            )
            if not any(values):
                continue
            unique_rows.setdefault(dedupe_key, values)
            record_count += 1

    if json_row_count > 0 and missing_records_count == json_row_count:
        raise ValueError(
            "No records were resolved from any JSON row. "
            "Check the real JSON structure and records-path."
        )

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "extracted"
    output_worksheet.append(args.fields)
    for values in unique_rows.values():
        output_worksheet.append(values)
    output_workbook.save(output_path)

    print(f"SOURCE={input_path}")
    print(f"SHEET={sheet_name}")
    print(f"JSON_COLUMN={args.json_column}")
    print(f"RECORDS_PATH={args.records_path or '<root>'}")
    print(f"OUTPUT={output_path}")
    print(f"ROWS={row_count}")
    print(f"JSON_ROWS={json_row_count}")
    print(f"RECORDS={record_count}")
    print(f"DEDUPED_ROWS={len(unique_rows)}")
    print(f"INVALID_JSON={invalid_json_count}")
    print(f"MISSING_RECORDS={missing_records_count}")


if __name__ == "__main__":
    main()
